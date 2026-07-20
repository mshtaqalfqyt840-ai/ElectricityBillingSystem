import io
import os
from django.conf import settings
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import arabic_reshaper
from bidi.algorithm import get_display

from .services.financial_report_service import FinancialReportService
from .services.audit_service import AuditLogger
from .permissions import IsAuthenticatedCustom, IsAdminOrAccountant
from .exceptions import ValidationError

class FinancialReportView(APIView):
    permission_classes = [IsAuthenticatedCustom, IsAdminOrAccountant]

    def get(self, request):
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        building_id = request.query_params.get('building_id')
        fmt = request.query_params.get('format', 'json').lower()

        if not start_date or not end_date:
            raise ValidationError("تاريخ البداية والنهاية (start_date, end_date) مطلوبان.")

        try:
            report_data = FinancialReportService.generate(start_date, end_date, building_id)
        except ValueError as e:
            raise ValidationError(str(e))

        AuditLogger.log(
            actor=request.user,
            action_type='financial_report_generated',
            target_model='FinancialReport',
            target_id=f"{start_date}_to_{end_date}",
            description=f"إنشاء تقرير مالي بصيغة {fmt} للفترة من {start_date} إلى {end_date}",
            metadata={
                'start_date': start_date,
                'end_date': end_date,
                'building_id': building_id,
                'format': fmt
            }
        )

        if fmt == 'excel':
            return self._generate_excel(report_data)
        elif fmt == 'pdf':
            return self._generate_pdf(report_data)
        else:
            return Response({"success": True, "data": report_data})

    def _generate_excel(self, data):
        wb = openpyxl.Workbook()
        
        # 1. الملخص العام
        ws_summary = wb.active
        ws_summary.title = "الملخص العام"
        ws_summary.sheet_view.rightToLeft = True

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

        headers = ["البيان", "القيمة"]
        ws_summary.append(headers)
        for cell in ws_summary[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        s = data['summary']
        ws_summary.append(["الفترة من", data['period']['start']])
        ws_summary.append(["الفترة إلى", data['period']['end']])
        ws_summary.append(["إجمالي الاستهلاك (كيلوواط)", s['total_consumption']])
        ws_summary.append(["إجمالي المبالغ المحصلة (ريال)", s['total_collected']])
        ws_summary.append(["إجمالي المديونيات المعلقة (ريال)", s['total_outstanding']])
        ws_summary.append(["عدد الفواتير الصادرة", s['issued_count']])
        ws_summary.append(["عدد الفواتير المسددة", s['paid_count']])
        ws_summary.append(["عدد الفواتير المتأخرة", s['overdue_count']])

        for row in ws_summary.iter_rows(min_row=2, max_col=2):
            for cell in row:
                cell.alignment = Alignment(horizontal='right')

        ws_summary.column_dimensions['A'].width = 35
        ws_summary.column_dimensions['B'].width = 20

        # 2. التفصيل حسب المبنى
        if not data['building_id'] and data['buildings_detail']:
            ws_details = wb.create_sheet(title="التفصيل حسب المبنى")
            ws_details.sheet_view.rightToLeft = True

            det_headers = [
                "المبنى", "الاستهلاك", "المحصل", "المديونيات", 
                "فواتير صادرة", "مسددة", "متأخرة"
            ]
            ws_details.append(det_headers)
            for cell in ws_details[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')

            for b in data['buildings_detail']:
                ws_details.append([
                    b['building_name'],
                    b['total_consumption'],
                    b['total_collected'],
                    b['total_outstanding'],
                    b['issued_count'],
                    b['paid_count'],
                    b['overdue_count']
                ])

            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
                ws_details.column_dimensions[col].width = 18

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output, 
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="financial_report_{data["period"]["start"]}_to_{data["period"]["end"]}.xlsx"'
        return response

    def _generate_pdf(self, data):
        output = io.BytesIO()
        c = canvas.Canvas(output, pagesize=A4)
        width, height = A4

        # Register Font
        font_path = os.path.join(settings.BASE_DIR, 'assets', 'fonts', 'Cairo-Regular.ttf')
        try:
            pdfmetrics.registerFont(TTFont('Cairo', font_path))
            c.setFont("Cairo", 14)
        except Exception:
            c.setFont("Helvetica", 14)

        def arabic(text):
            return get_display(arabic_reshaper.reshape(str(text)))

        c.drawRightString(width - 50, height - 50, arabic("تقرير مالي - نظام الفواتير الكهربائية"))
        c.setFont("Cairo" if 'Cairo' in pdfmetrics.getRegisteredFontNames() else "Helvetica", 12)
        c.drawRightString(width - 50, height - 80, arabic(f"الفترة من {data['period']['start']} إلى {data['period']['end']}"))
        
        s = data['summary']
        y = height - 130
        
        c.drawRightString(width - 50, y, arabic("الملخص العام:"))
        y -= 30
        
        lines = [
            (arabic("إجمالي الاستهلاك:"), f"{s['total_consumption']} KWh"),
            (arabic("إجمالي المبالغ المحصلة:"), f"{s['total_collected']} SAR"),
            (arabic("إجمالي المديونيات المعلقة:"), f"{s['total_outstanding']} SAR"),
            (arabic("الفواتير (صادرة/مسددة/متأخرة):"), f"{s['issued_count']} / {s['paid_count']} / {s['overdue_count']}"),
        ]

        for label, val in lines:
            c.drawRightString(width - 80, y, label)
            c.drawString(width - 300, y, val)
            y -= 25

        if not data['building_id'] and data['buildings_detail']:
            y -= 20
            c.drawRightString(width - 50, y, arabic("التفصيل حسب المبنى:"))
            y -= 30

            # Headers
            c.drawRightString(width - 50, y, arabic("المبنى"))
            c.drawRightString(width - 150, y, arabic("الاستهلاك"))
            c.drawRightString(width - 250, y, arabic("المحصل"))
            c.drawRightString(width - 350, y, arabic("المديونيات"))
            y -= 20

            for b in data['buildings_detail']:
                if y < 50:
                    c.showPage()
                    c.setFont("Cairo" if 'Cairo' in pdfmetrics.getRegisteredFontNames() else "Helvetica", 12)
                    y = height - 50

                c.drawRightString(width - 50, y, arabic(b['building_name']))
                c.drawRightString(width - 150, y, str(b['total_consumption']))
                c.drawRightString(width - 250, y, str(b['total_collected']))
                c.drawRightString(width - 350, y, str(b['total_outstanding']))
                y -= 20

        c.save()
        output.seek(0)

        response = HttpResponse(output, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="financial_report_{data["period"]["start"]}_to_{data["period"]["end"]}.pdf"'
        return response
